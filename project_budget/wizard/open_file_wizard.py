from odoo import models, fields, _
from openpyxl import load_workbook
from io import BytesIO
import base64


class OpenFileWizard(models.TransientModel):
    _name = 'project_budget.open_file_wizard'
    _description = 'Open File Wizard'

    company_id = fields.Many2one('res.company', string='Company', required=True)
    data_file = fields.Binary(string='Data file', required=True)

    def import_file_data_from_wizard(self):
        self.import_file_data(BytesIO(base64.b64decode(self.data_file)))

    def import_file_data(self, file_name):
        wb = load_workbook(file_name, read_only=True, data_only=True)
        ws = wb.active
        data = []
        for record in ws.iter_rows(min_row=6, max_row=6, min_col=4, max_col=None, values_only=True):
            data = str(record).strip('()')
        if data:
            self.env['project_budget.report_external_data'].create({
                'company_id': self.company_id.id,
                'data': data,
                'file': self.data_file,
            })
